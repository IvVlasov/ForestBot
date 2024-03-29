from openpyxl import Workbook, load_workbook
import os
from settings import EXCEL_PATH


class ExcelFile:

    @classmethod
    def create_excel(cls):
        if os.path.exists(EXCEL_PATH):
            return
        wb = Workbook()
        ws = wb.active
        ws.title = 'Лист 1'
        ws['A1'] = 'Номер заявки'
        ws['B1'] = 'Нарушение'
        ws['C1'] = 'Место'
        ws['D1'] = 'Описание'
        ws['E1'] = 'Контакт'
        ws['F1'] = 'Дополнительный контакт'
        ws['G1'] = 'Файлы'
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 100
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        wb.save(filename=EXCEL_PATH)

    @classmethod
    def paste_in_excel(cls, data):
        cls.create_excel()
        wb = Workbook()
        wb = load_workbook(filename=EXCEL_PATH)
        ws = wb.active
        idx = str(ws.max_row + 1)
        ws['A'+idx] = data['record_id']
        ws['B'+idx] = data['type']
        ws['C'+idx] = data['place']
        ws['D'+idx] = data['descr']
        ws['E'+idx] = data.get('phone', '-')
        ws['F'+idx] = data.get('contact', '-')
        ws['G'+idx] = str(len(data['files_list']))
        wb.save(filename=EXCEL_PATH)
